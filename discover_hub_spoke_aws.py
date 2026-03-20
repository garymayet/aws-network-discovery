#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
╔══════════════════════════════════════════════════════════════════════════════╗
║  Discover-HubSpoke-AWS.py                                                  ║
║  Descubrimiento profundo de arquitectura Hub & Spoke en AWS                ║
║                                                                            ║
║  Equivalente funcional de Discover-HubSpoke.ps1 (Azure) para entornos AWS  ║
║                                                                            ║
║  Pilares:                                                                  ║
║    1. Descubrimiento Multi-Cuenta / Multi-Región                           ║
║    2. Clasificación Hub vs Spoke (heurística automática)                   ║
║    3. Evaluación de Mejores Prácticas (PASS/FAIL/WARNING → CSV)            ║
║    4. Generación de Diagrama Mermaid.js                                    ║
║    5. Exportación: CSVs + .mmd + Markdown Summary                          ║
║                                                                            ║
║  Requisitos: Python 3.9+, boto3, botocore                                  ║
║  Autor: Arquitecto de Soluciones Cloud Senior                              ║
║  Fecha: 2026-03-18                                                         ║
╚══════════════════════════════════════════════════════════════════════════════╝

Modos de autenticación soportados:
  1. --profiles perfil1,perfil2    → Usa perfiles de AWS CLI (~/.aws/credentials)
  2. --role-arn arn:aws:iam::*:role/X  → Asume un rol cross-account en cada cuenta
  3. --org                         → Usa AWS Organizations para listar todas las cuentas
  4. (sin flags)                   → Usa las credenciales por defecto del entorno

Uso:
  python3 discover_hub_spoke_aws.py --regions us-east-1,us-west-2
  python3 discover_hub_spoke_aws.py --profiles prod,dev --regions us-east-1
  python3 discover_hub_spoke_aws.py --org --role-arn arn:aws:iam::{account_id}:role/OrgReadOnly
  python3 discover_hub_spoke_aws.py --account-ids 111111111111,222222222222 \\
                                    --role-arn arn:aws:iam::{account_id}:role/NetAudit
"""

import argparse
import csv
import json
import os
import re
import sys
import traceback
from collections import defaultdict
from datetime import datetime, timezone
from typing import Any, Optional

try:
    import boto3
    import botocore
    from botocore.config import Config as BotoConfig
    from botocore.exceptions import (
        ClientError,
        NoCredentialsError,
        ProfileNotFound,
        EndpointConnectionError,
    )
except ImportError:
    print("[ERROR] boto3 no está instalado. Ejecutar: pip install boto3")
    sys.exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers,
    )
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ─────────────────────────────────────────────────────────────────────────────
# CONSTANTES
# ─────────────────────────────────────────────────────────────────────────────

SCRIPT_VERSION = "1.1.0"
DEFAULT_REGIONS = [
    "us-east-1", "us-east-2", "us-west-1", "us-west-2",
    "eu-west-1", "eu-central-1", "sa-east-1",
]
BOTO_CONFIG = BotoConfig(
    retries={"max_attempts": 5, "mode": "adaptive"},
    connect_timeout=10,
    read_timeout=30,
)
TIMESTAMP = datetime.now(timezone.utc).strftime("%Y-%m-%d_%H%M%S")


# ─────────────────────────────────────────────────────────────────────────────
# FUNCIONES AUXILIARES — Consola con color
# ─────────────────────────────────────────────────────────────────────────────

class Colors:
    """ANSI escape codes para salida en consola."""
    CYAN    = "\033[96m"
    GREEN   = "\033[92m"
    YELLOW  = "\033[93m"
    RED     = "\033[91m"
    BOLD    = "\033[1m"
    RESET   = "\033[0m"
    MAGENTA = "\033[95m"


def write_status(message: str, level: str = "INFO") -> None:
    """Imprime un mensaje con formato y código de color según el nivel."""
    color_map = {
        "INFO":  Colors.CYAN,
        "OK":    Colors.GREEN,
        "WARN":  Colors.YELLOW,
        "ERROR": Colors.RED,
    }
    ts = datetime.now().strftime("%H:%M:%S")
    color = color_map.get(level, Colors.RESET)
    print(f"{color}[{ts}] [{level:5s}] {message}{Colors.RESET}")


def sanitize_mermaid_id(name: str) -> str:
    """Convierte un nombre arbitrario en un ID válido para Mermaid.js."""
    return re.sub(r"[^a-zA-Z0-9]", "_", name)


def paginate(client, method: str, key: str, **kwargs) -> list:
    """
    Helper genérico de paginación para llamadas Boto3.
    Maneja automáticamente NextToken / Marker según el servicio.
    """
    results = []
    paginator = client.get_paginator(method)
    try:
        for page in paginator.paginate(**kwargs):
            results.extend(page.get(key, []))
    except ClientError as e:
        code = e.response["Error"]["Code"]
        if code in ("AccessDeniedException", "UnauthorizedAccess", "AuthorizationError"):
            write_status(f"  Sin permisos para {method}: {code}", "WARN")
        else:
            write_status(f"  Error en {method}: {e}", "ERROR")
    except Exception as e:
        write_status(f"  Error inesperado en {method}: {e}", "WARN")
    return results


def safe_api_call(func, *args, default=None, **kwargs):
    """
    Envuelve una llamada a la API de AWS con manejo de errores estándar.
    Retorna el resultado o el valor por defecto si falla.
    """
    try:
        return func(*args, **kwargs)
    except ClientError as e:
        code = e.response["Error"]["Code"]
        if code in ("AccessDeniedException", "UnauthorizedAccess", "AuthorizationError"):
            write_status(f"  Sin permisos: {code}", "WARN")
        elif code == "ResourceNotFoundException":
            pass  # recurso no existe, silenciar
        else:
            write_status(f"  ClientError: {e}", "WARN")
    except EndpointConnectionError:
        pass  # servicio no disponible en esta región
    except Exception as e:
        write_status(f"  Error inesperado: {e}", "WARN")
    return default


# ─────────────────────────────────────────────────────────────────────────────
# ESTRUCTURAS DE DATOS (Data Store)
# ─────────────────────────────────────────────────────────────────────────────

class DiscoveryStore:
    """Almacena todos los recursos descubiertos durante la exploración."""

    def __init__(self):
        self.vpcs: list[dict] = []
        self.subnets: list[dict] = []
        self.route_tables: list[dict] = []
        self.tgws: list[dict] = []
        self.tgw_attachments: list[dict] = []
        self.vpc_peerings: list[dict] = []
        self.vgws: list[dict] = []
        self.cgws: list[dict] = []
        self.vpn_connections: list[dict] = []
        self.dx_connections: list[dict] = []
        self.network_firewalls: list[dict] = []
        self.nfw_endpoints: list[dict] = []
        self.r53_resolver_endpoints: list[dict] = []
        self.nat_gateways: list[dict] = []
        self.best_practices: list[dict] = []

    def summary(self) -> dict:
        return {
            "VPCs": len(self.vpcs),
            "Subnets": len(self.subnets),
            "Route Tables": len(self.route_tables),
            "Transit Gateways": len(self.tgws),
            "TGW Attachments": len(self.tgw_attachments),
            "VPC Peerings": len(self.vpc_peerings),
            "Virtual Private GWs": len(self.vgws),
            "Customer GWs": len(self.cgws),
            "VPN Connections": len(self.vpn_connections),
            "Direct Connect": len(self.dx_connections),
            "Network Firewalls": len(self.network_firewalls),
            "R53 Resolver Endpoints": len(self.r53_resolver_endpoints),
            "NAT Gateways": len(self.nat_gateways),
        }


# ─────────────────────────────────────────────────────────────────────────────
# RESOLUCIÓN DE CUENTAS Y SESIONES
# ─────────────────────────────────────────────────────────────────────────────

def resolve_accounts(args) -> list[dict]:
    """
    Resuelve la lista de cuentas AWS a auditar según los argumentos.
    Retorna una lista de dicts con 'account_id', 'label' y 'session'.
    """
    accounts = []

    # ── Modo 1: Perfiles explícitos ──
    if args.profiles:
        for profile in args.profiles.split(","):
            profile = profile.strip()
            try:
                session = boto3.Session(profile_name=profile)
                sts = session.client("sts", config=BOTO_CONFIG)
                identity = sts.get_caller_identity()
                accounts.append({
                    "account_id": identity["Account"],
                    "label": f"profile:{profile}",
                    "session": session,
                })
                write_status(f"  Perfil '{profile}' → Cuenta {identity['Account']}", "OK")
            except (ProfileNotFound, NoCredentialsError, ClientError) as e:
                write_status(f"  Error con perfil '{profile}': {e}", "ERROR")
        return accounts

    # ── Modo 2: AWS Organizations ──
    if args.org:
        session = boto3.Session()
        org_client = session.client("organizations", config=BOTO_CONFIG)
        try:
            org_accounts = paginate(org_client, "list_accounts", "Accounts")
            active = [a for a in org_accounts if a["Status"] == "ACTIVE"]
            write_status(f"  AWS Organizations: {len(active)} cuentas activas encontradas", "OK")
            for acct in active:
                acct_id = acct["Id"]
                acct_session = _assume_role_session(session, acct_id, args.role_arn)
                if acct_session:
                    accounts.append({
                        "account_id": acct_id,
                        "label": f"org:{acct.get('Name', acct_id)}",
                        "session": acct_session,
                    })
        except ClientError as e:
            write_status(f"  Error al listar cuentas de Organizations: {e}", "ERROR")
        return accounts

    # ── Modo 3: Lista explícita de account IDs + AssumeRole ──
    if args.account_ids:
        session = boto3.Session()
        for acct_id in args.account_ids.split(","):
            acct_id = acct_id.strip()
            acct_session = _assume_role_session(session, acct_id, args.role_arn)
            if acct_session:
                accounts.append({
                    "account_id": acct_id,
                    "label": f"cross-account:{acct_id}",
                    "session": acct_session,
                })
        return accounts

    # ── Modo 4: Credenciales por defecto (cuenta única) ──
    try:
        session = boto3.Session()
        sts = session.client("sts", config=BOTO_CONFIG)
        identity = sts.get_caller_identity()
        accounts.append({
            "account_id": identity["Account"],
            "label": f"default:{identity['Account']}",
            "session": session,
        })
        write_status(f"  Usando credenciales por defecto → Cuenta {identity['Account']}", "OK")
    except (NoCredentialsError, ClientError) as e:
        write_status(f"  No se encontraron credenciales AWS: {e}", "ERROR")
        sys.exit(1)

    return accounts


def _assume_role_session(
    base_session: boto3.Session,
    account_id: str,
    role_arn_template: Optional[str],
) -> Optional[boto3.Session]:
    """
    Asume un rol cross-account y retorna una sesión temporal.
    El template puede contener {account_id} como placeholder.
    """
    if not role_arn_template:
        write_status(
            f"  Cuenta {account_id}: Se requiere --role-arn para cross-account", "WARN"
        )
        return None

    role_arn = role_arn_template.replace("{account_id}", account_id)
    try:
        sts = base_session.client("sts", config=BOTO_CONFIG)
        creds = sts.assume_role(
            RoleArn=role_arn,
            RoleSessionName=f"HubSpokeDiscovery-{account_id}",
            DurationSeconds=3600,
        )["Credentials"]
        session = boto3.Session(
            aws_access_key_id=creds["AccessKeyId"],
            aws_secret_access_key=creds["SecretAccessKey"],
            aws_session_token=creds["SessionToken"],
        )
        write_status(f"  Rol asumido en cuenta {account_id}: {role_arn}", "OK")
        return session
    except ClientError as e:
        write_status(f"  No se pudo asumir rol en {account_id}: {e}", "ERROR")
        return None


# ─────────────────────────────────────────────────────────────────────────────
# PILAR 1 — RECOLECCIÓN DE RECURSOS POR CUENTA / REGIÓN
# ─────────────────────────────────────────────────────────────────────────────

def collect_resources(
    session: boto3.Session,
    account_id: str,
    account_label: str,
    regions: list[str],
    store: DiscoveryStore,
) -> None:
    """Itera regiones y recolecta todos los recursos de red relevantes."""

    for region in regions:
        write_status(f"  Región: {region}")
        ec2 = session.client("ec2", region_name=region, config=BOTO_CONFIG)

        # ── 1.1 VPCs ─────────────────────────────────────────────────────
        write_status(f"    Descubriendo VPCs...")
        vpcs_raw = safe_api_call(
            lambda: paginate(ec2, "describe_vpcs", "Vpcs"), default=[]
        )
        for vpc in vpcs_raw:
            vpc_id = vpc["VpcId"]
            name_tag = _get_name_tag(vpc.get("Tags", []))
            store.vpcs.append({
                "AccountId": account_id,
                "AccountLabel": account_label,
                "Region": region,
                "VpcId": vpc_id,
                "VpcName": name_tag or vpc_id,
                "CidrBlock": vpc.get("CidrBlock", ""),
                "AdditionalCidrs": ", ".join(
                    [a["CidrBlock"] for a in vpc.get("CidrBlockAssociationSet", [])
                     if a.get("CidrBlockState", {}).get("State") == "associated"
                     and a["CidrBlock"] != vpc.get("CidrBlock")]
                ),
                "IsDefault": vpc.get("IsDefault", False),
                "State": vpc.get("State", ""),
                # Clasificación (se determina en Pilar 2)
                "IsHub": False,
                "IsSpoke": False,
                "HubReasons": [],
            })

        # ── 1.2 Subnets ──────────────────────────────────────────────────
        write_status(f"    Descubriendo Subnets...")
        subnets_raw = safe_api_call(
            lambda: paginate(ec2, "describe_subnets", "Subnets"), default=[]
        )
        for sn in subnets_raw:
            name_tag = _get_name_tag(sn.get("Tags", []))
            store.subnets.append({
                "AccountId": account_id,
                "Region": region,
                "VpcId": sn["VpcId"],
                "SubnetId": sn["SubnetId"],
                "SubnetName": name_tag or sn["SubnetId"],
                "CidrBlock": sn.get("CidrBlock", ""),
                "AvailabilityZone": sn.get("AvailabilityZone", ""),
                "MapPublicIp": sn.get("MapPublicIpOnLaunch", False),
            })

        # ── 1.3 Route Tables ─────────────────────────────────────────────
        write_status(f"    Descubriendo Route Tables...")
        rts_raw = safe_api_call(
            lambda: paginate(ec2, "describe_route_tables", "RouteTables"), default=[]
        )
        for rt in rts_raw:
            name_tag = _get_name_tag(rt.get("Tags", []))
            associations = rt.get("Associations", [])
            is_main = any(a.get("Main", False) for a in associations)
            assoc_subnets = [a.get("SubnetId", "") for a in associations if a.get("SubnetId")]
            # Analizar rutas para detectar targets TGW / NAT / VGW / NFW
            route_targets = []
            for route in rt.get("Routes", []):
                dest = route.get("DestinationCidrBlock", route.get("DestinationPrefixListId", ""))
                gw_id = route.get("GatewayId", "")
                nat_id = route.get("NatGatewayId", "")
                tgw_id = route.get("TransitGatewayId", "")
                vpc_ep = route.get("VpcPeeringConnectionId", "")
                nif_id = route.get("NetworkInterfaceId", "")
                target = tgw_id or nat_id or gw_id or vpc_ep or nif_id or "local"
                route_targets.append(f"{dest}→{target}")

            store.route_tables.append({
                "AccountId": account_id,
                "Region": region,
                "VpcId": rt.get("VpcId", ""),
                "RouteTableId": rt["RouteTableId"],
                "RouteTableName": name_tag or rt["RouteTableId"],
                "IsMain": is_main,
                "AssociatedSubnets": ", ".join(assoc_subnets),
                "Routes": "; ".join(route_targets[:20]),  # limitar para CSV
            })

        # ── 1.4 Transit Gateways ─────────────────────────────────────────
        write_status(f"    Descubriendo Transit Gateways...")
        tgws_raw = safe_api_call(
            lambda: paginate(ec2, "describe_transit_gateways", "TransitGateways"),
            default=[],
        )
        for tgw in tgws_raw:
            name_tag = _get_name_tag(tgw.get("Tags", []))
            opts = tgw.get("Options", {})
            store.tgws.append({
                "AccountId": account_id,
                "Region": region,
                "TgwId": tgw["TransitGatewayId"],
                "TgwName": name_tag or tgw["TransitGatewayId"],
                "State": tgw.get("State", ""),
                "OwnerId": tgw.get("OwnerId", ""),
                "AmazonSideAsn": opts.get("AmazonSideAsn", ""),
                "AutoAcceptSharedAttachments": opts.get("AutoAcceptSharedAttachments", ""),
                "DefaultRouteTableAssociation": opts.get("DefaultRouteTableAssociation", ""),
                "DefaultRouteTablePropagation": opts.get("DefaultRouteTablePropagation", ""),
            })

        # ── 1.5 TGW Attachments ──────────────────────────────────────────
        write_status(f"    Descubriendo TGW Attachments...")
        tgw_att_raw = safe_api_call(
            lambda: paginate(
                ec2, "describe_transit_gateway_attachments", "TransitGatewayAttachments"
            ),
            default=[],
        )
        for att in tgw_att_raw:
            name_tag = _get_name_tag(att.get("Tags", []))
            store.tgw_attachments.append({
                "AccountId": account_id,
                "Region": region,
                "TgwAttachmentId": att["TransitGatewayAttachmentId"],
                "TgwAttachmentName": name_tag or att["TransitGatewayAttachmentId"],
                "TgwId": att.get("TransitGatewayId", ""),
                "ResourceType": att.get("ResourceType", ""),  # vpc, vpn, direct-connect-gateway, peering, tgw-peering
                "ResourceId": att.get("ResourceId", ""),
                "ResourceOwnerId": att.get("ResourceOwnerId", ""),
                "State": att.get("State", ""),
                "Association": att.get("Association", {}).get("TransitGatewayRouteTableId", ""),
            })

        # Obtener detalle de VPC attachments (AZs + Appliance Mode)
        for att in store.tgw_attachments:
            if att["ResourceType"] == "vpc" and att["Region"] == region and att["AccountId"] == account_id:
                _enrich_tgw_vpc_attachment(ec2, att)

        # ── 1.6 VPC Peerings ─────────────────────────────────────────────
        write_status(f"    Descubriendo VPC Peerings...")
        peerings_raw = safe_api_call(
            lambda: paginate(
                ec2, "describe_vpc_peering_connections", "VpcPeeringConnections"
            ),
            default=[],
        )
        for pcx in peerings_raw:
            name_tag = _get_name_tag(pcx.get("Tags", []))
            req = pcx.get("RequesterVpcInfo", {})
            acc = pcx.get("AccepterVpcInfo", {})
            store.vpc_peerings.append({
                "AccountId": account_id,
                "Region": region,
                "PeeringId": pcx["VpcPeeringConnectionId"],
                "PeeringName": name_tag or pcx["VpcPeeringConnectionId"],
                "Status": pcx.get("Status", {}).get("Code", ""),
                "RequesterVpcId": req.get("VpcId", ""),
                "RequesterOwnerId": req.get("OwnerId", ""),
                "RequesterCidr": req.get("CidrBlock", ""),
                "AccepterVpcId": acc.get("VpcId", ""),
                "AccepterOwnerId": acc.get("OwnerId", ""),
                "AccepterCidr": acc.get("CidrBlock", ""),
            })

        # ── 1.7 Virtual Private Gateways ─────────────────────────────────
        write_status(f"    Descubriendo VPN / VGWs...")
        vgws_raw = safe_api_call(
            lambda: ec2.describe_vpn_gateways().get("VpnGateways", []), default=[]
        )
        for vgw in vgws_raw:
            name_tag = _get_name_tag(vgw.get("Tags", []))
            attached_vpcs = [
                a["VpcId"] for a in vgw.get("VpcAttachments", [])
                if a.get("State") == "attached"
            ]
            store.vgws.append({
                "AccountId": account_id,
                "Region": region,
                "VgwId": vgw["VpnGatewayId"],
                "VgwName": name_tag or vgw["VpnGatewayId"],
                "State": vgw.get("State", ""),
                "Type": vgw.get("Type", ""),
                "AmazonSideAsn": vgw.get("AmazonSideAsn", ""),
                "AttachedVpcs": ", ".join(attached_vpcs),
                "AvailabilityZone": vgw.get("AvailabilityZone", ""),
            })

        # ── 1.8 Customer Gateways ────────────────────────────────────────
        cgws_raw = safe_api_call(
            lambda: ec2.describe_customer_gateways().get("CustomerGateways", []),
            default=[],
        )
        for cgw in cgws_raw:
            name_tag = _get_name_tag(cgw.get("Tags", []))
            store.cgws.append({
                "AccountId": account_id,
                "Region": region,
                "CgwId": cgw["CustomerGatewayId"],
                "CgwName": name_tag or cgw["CustomerGatewayId"],
                "BgpAsn": cgw.get("BgpAsn", ""),
                "IpAddress": cgw.get("IpAddress", ""),
                "State": cgw.get("State", ""),
                "Type": cgw.get("Type", ""),
            })

        # ── 1.9 VPN Connections ──────────────────────────────────────────
        vpns_raw = safe_api_call(
            lambda: ec2.describe_vpn_connections().get("VpnConnections", []),
            default=[],
        )
        for vpn in vpns_raw:
            name_tag = _get_name_tag(vpn.get("Tags", []))
            tunnels = vpn.get("VgwTelemetry", [])
            tunnel_ips = [t.get("OutsideIpAddress", "") for t in tunnels]
            tunnel_statuses = [t.get("Status", "") for t in tunnels]
            store.vpn_connections.append({
                "AccountId": account_id,
                "Region": region,
                "VpnId": vpn["VpnConnectionId"],
                "VpnName": name_tag or vpn["VpnConnectionId"],
                "State": vpn.get("State", ""),
                "Type": vpn.get("Type", ""),
                "VgwId": vpn.get("VpnGatewayId", ""),
                "TgwId": vpn.get("TransitGatewayId", ""),
                "CgwId": vpn.get("CustomerGatewayId", ""),
                "Category": vpn.get("Category", ""),
                "StaticRoutesOnly": vpn.get("Options", {}).get("StaticRoutesOnly", False),
                "TunnelIPs": ", ".join(tunnel_ips),
                "TunnelStatuses": ", ".join(tunnel_statuses),
            })

        # ── 1.10 Direct Connect Connections ──────────────────────────────
        write_status(f"    Descubriendo Direct Connect...")
        try:
            dx = session.client("directconnect", region_name=region, config=BOTO_CONFIG)
            dx_raw = safe_api_call(
                lambda: dx.describe_connections().get("connections", []),
                default=[],
            )
            for conn in dx_raw:
                store.dx_connections.append({
                    "AccountId": account_id,
                    "Region": region,
                    "ConnectionId": conn.get("connectionId", ""),
                    "ConnectionName": conn.get("connectionName", ""),
                    "State": conn.get("connectionState", ""),
                    "Bandwidth": conn.get("bandwidth", ""),
                    "Location": conn.get("location", ""),
                    "PartnerName": conn.get("partnerName", ""),
                    "Vlan": conn.get("vlan", ""),
                    "AwsDevice": conn.get("awsDevice", ""),
                    "HasLogicalRedundancy": conn.get("hasLogicalRedundancy", ""),
                })
        except Exception:
            pass  # DX no disponible en esta región o sin permisos

        # ── 1.11 AWS Network Firewall ────────────────────────────────────
        write_status(f"    Descubriendo Network Firewalls...")
        try:
            nfw = session.client(
                "network-firewall", region_name=region, config=BOTO_CONFIG
            )
            fw_list = safe_api_call(
                lambda: paginate(nfw, "list_firewalls", "Firewalls"), default=[]
            )
            for fw_summary in fw_list:
                fw_arn = fw_summary.get("FirewallArn", "")
                fw_name = fw_summary.get("FirewallName", "")
                # Obtener detalle completo del firewall
                fw_detail = safe_api_call(
                    lambda: nfw.describe_firewall(FirewallArn=fw_arn),
                    default={},
                )
                fw = fw_detail.get("Firewall", {}) if fw_detail else {}
                fw_status = fw_detail.get("FirewallStatus", {}) if fw_detail else {}

                vpc_id = fw.get("VpcId", "")
                subnet_mappings = fw.get("SubnetMappings", [])
                subnet_ids = [sm.get("SubnetId", "") for sm in subnet_mappings]

                # Endpoints del firewall (para mapeo a AZs)
                sync_states = fw_status.get("SyncStates", {})
                endpoint_azs = list(sync_states.keys()) if sync_states else []
                endpoint_ids = []
                for az, state in sync_states.items():
                    att = state.get("Attachment", {})
                    ep_id = att.get("EndpointId", "")
                    if ep_id:
                        endpoint_ids.append(ep_id)
                        store.nfw_endpoints.append({
                            "AccountId": account_id,
                            "Region": region,
                            "FirewallName": fw_name,
                            "VpcId": vpc_id,
                            "AZ": az,
                            "EndpointId": ep_id,
                            "SubnetId": att.get("SubnetId", ""),
                            "Status": att.get("Status", ""),
                        })

                store.network_firewalls.append({
                    "AccountId": account_id,
                    "Region": region,
                    "FirewallName": fw_name,
                    "FirewallArn": fw_arn,
                    "VpcId": vpc_id,
                    "SubnetIds": ", ".join(subnet_ids),
                    "EndpointAZs": ", ".join(endpoint_azs),
                    "EndpointCount": len(endpoint_ids),
                    "PolicyArn": fw.get("FirewallPolicyArn", ""),
                    "DeleteProtection": fw.get("DeleteProtection", False),
                    "Description": fw.get("Description", ""),
                })
        except Exception:
            pass  # Network Firewall no disponible

        # ── 1.12 Route 53 Resolver Endpoints ─────────────────────────────
        write_status(f"    Descubriendo Route 53 Resolver Endpoints...")
        try:
            r53r = session.client(
                "route53resolver", region_name=region, config=BOTO_CONFIG
            )
            resolver_eps = safe_api_call(
                lambda: paginate(
                    r53r, "list_resolver_endpoints", "ResolverEndpoints"
                ),
                default=[],
            )
            for ep in resolver_eps:
                ep_id = ep.get("Id", "")
                # Obtener IPs del endpoint para determinar AZs
                ip_addrs = safe_api_call(
                    lambda: paginate(
                        r53r,
                        "list_resolver_endpoint_ip_addresses",
                        "IpAddresses",
                        ResolverEndpointId=ep_id,
                    ),
                    default=[],
                )
                ep_azs = list(set(ip.get("SubnetId", "")[:ip.get("SubnetId", "").rfind("-")] for ip in ip_addrs)) if ip_addrs else []
                subnet_ids = [ip.get("SubnetId", "") for ip in ip_addrs]

                # Determinar el VpcId (los Resolver Endpoints no lo reportan directamente;
                # la forma canónica es que el HostVPCId viene en el endpoint).
                host_vpc = ep.get("HostVPCId", "")

                store.r53_resolver_endpoints.append({
                    "AccountId": account_id,
                    "Region": region,
                    "EndpointId": ep_id,
                    "EndpointName": ep.get("Name", ep_id),
                    "Direction": ep.get("Direction", ""),  # INBOUND / OUTBOUND
                    "VpcId": host_vpc,
                    "Status": ep.get("Status", ""),
                    "SubnetIds": ", ".join(subnet_ids),
                    "IpAddressCount": ep.get("IpAddressCount", 0),
                    "SecurityGroupIds": ", ".join(ep.get("SecurityGroupIds", [])),
                })
        except Exception:
            pass

        # ── 1.13 NAT Gateways ────────────────────────────────────────────
        write_status(f"    Descubriendo NAT Gateways...")
        nats_raw = safe_api_call(
            lambda: paginate(ec2, "describe_nat_gateways", "NatGateways"), default=[]
        )
        for nat in nats_raw:
            if nat.get("State") == "deleted":
                continue
            name_tag = _get_name_tag(nat.get("Tags", []))
            pub_ips = [
                addr.get("PublicIp", "")
                for addr in nat.get("NatGatewayAddresses", [])
                if addr.get("PublicIp")
            ]
            store.nat_gateways.append({
                "AccountId": account_id,
                "Region": region,
                "NatGatewayId": nat["NatGatewayId"],
                "NatGatewayName": name_tag or nat["NatGatewayId"],
                "VpcId": nat.get("VpcId", ""),
                "SubnetId": nat.get("SubnetId", ""),
                "State": nat.get("State", ""),
                "ConnectivityType": nat.get("ConnectivityType", "public"),
                "PublicIPs": ", ".join(pub_ips),
            })

    write_status(
        f"  Recolección finalizada para cuenta {account_id}", "OK"
    )


def _enrich_tgw_vpc_attachment(ec2_client, attachment: dict) -> None:
    """Enriquece un TGW VPC attachment con detalles de AZs y Appliance Mode."""
    try:
        resp = ec2_client.describe_transit_gateway_vpc_attachments(
            TransitGatewayAttachmentIds=[attachment["TgwAttachmentId"]]
        )
        details = resp.get("TransitGatewayVpcAttachments", [])
        if details:
            det = details[0]
            subnet_ids = det.get("SubnetIds", [])
            attachment["VpcAttachmentSubnets"] = ", ".join(subnet_ids)
            attachment["VpcAttachmentAZCount"] = len(subnet_ids)
            opts = det.get("Options", {})
            attachment["ApplianceModeSupport"] = opts.get("ApplianceModeSupport", "disable")
            attachment["DnsSupport"] = opts.get("DnsSupport", "enable")
            attachment["Ipv6Support"] = opts.get("Ipv6Support", "disable")
    except Exception:
        attachment["VpcAttachmentSubnets"] = ""
        attachment["VpcAttachmentAZCount"] = 0
        attachment["ApplianceModeSupport"] = "unknown"


def _get_name_tag(tags: list[dict]) -> str:
    """Extrae el valor de la etiqueta 'Name' de una lista de tags de AWS."""
    for tag in tags:
        if tag.get("Key") == "Name":
            return tag.get("Value", "")
    return ""


# ─────────────────────────────────────────────────────────────────────────────
# PILAR 2 — CLASIFICACIÓN HUB vs SPOKE (Heurística Automática)
# ─────────────────────────────────────────────────────────────────────────────

def classify_hub_spoke(store: DiscoveryStore) -> None:
    """
    Clasifica cada VPC como Hub o Spoke según heurísticas de infraestructura.

    Criterios para Hub VPC:
      H1. Contiene un AWS Network Firewall.
      H2. Es el VPC de un TGW attachment donde se concentran múltiples attachments
          (actúa como VPC de inspección / tránsito central).
      H3. Concentra NAT Gateways de múltiples AZs (centralización de salida a internet).
      H4. Aloja endpoints de Route 53 Resolver (Inbound y/o Outbound).
      H5. Tiene un Virtual Private Gateway (VGW) attached (conectividad on-premises).

    Todo VPC que no sea Hub y tenga conectividad hacia un Hub se clasifica como Spoke.
    """
    write_status("Clasificando topología Hub vs Spoke...")

    hub_vpc_ids: dict[str, list[str]] = {}  # vpc_id -> lista de razones

    # ── H1: VPCs con Network Firewall ──
    for nfw in store.network_firewalls:
        vpc_id = nfw["VpcId"]
        if vpc_id:
            hub_vpc_ids.setdefault(vpc_id, []).append(
                f"Network Firewall: {nfw['FirewallName']}"
            )

    # ── H2: VPCs que son punto de tránsito central ──
    # Contar cuántos TGW attachments tipo "vpc" apuntan a cada TGW
    tgw_vpc_count = defaultdict(int)
    tgw_to_vpcs = defaultdict(set)
    for att in store.tgw_attachments:
        if att["ResourceType"] == "vpc":
            tgw_vpc_count[att["TgwId"]] += 1
            tgw_to_vpcs[att["TgwId"]].add(att["ResourceId"])

    # Si un TGW tiene >=2 VPC attachments, el que tiene NFW o más NATs es hub
    # También consideramos VPC de inspección: el que tiene Appliance Mode habilitado
    for att in store.tgw_attachments:
        if att["ResourceType"] == "vpc":
            if att.get("ApplianceModeSupport") == "enable":
                vpc_id = att["ResourceId"]
                hub_vpc_ids.setdefault(vpc_id, []).append(
                    f"TGW Appliance Mode habilitado ({att['TgwId']})"
                )

    # ── H3: VPCs con NAT Gateways concentrados ──
    nat_by_vpc = defaultdict(list)
    for nat in store.nat_gateways:
        if nat["State"] == "available":
            nat_by_vpc[nat["VpcId"]].append(nat)

    for vpc_id, nats in nat_by_vpc.items():
        if len(nats) >= 2:  # ≥2 NATs sugiere centralización
            azs = set()
            for nat in nats:
                # Buscar el AZ de la subnet del NAT
                for sn in store.subnets:
                    if sn["SubnetId"] == nat["SubnetId"]:
                        azs.add(sn["AvailabilityZone"])
                        break
            if len(azs) >= 2:
                hub_vpc_ids.setdefault(vpc_id, []).append(
                    f"NAT Gateways centralizados ({len(nats)} NATs en {len(azs)} AZs)"
                )

    # ── H4: VPCs con Route 53 Resolver Endpoints ──
    for ep in store.r53_resolver_endpoints:
        vpc_id = ep["VpcId"]
        if vpc_id:
            hub_vpc_ids.setdefault(vpc_id, []).append(
                f"Route 53 Resolver {ep['Direction']}: {ep['EndpointName']}"
            )

    # ── H5: VPCs con VGW (conectividad on-premises) ──
    for vgw in store.vgws:
        for vpc_id in vgw["AttachedVpcs"].split(", "):
            vpc_id = vpc_id.strip()
            if vpc_id:
                hub_vpc_ids.setdefault(vpc_id, []).append(
                    f"VGW attached: {vgw['VgwName']}"
                )

    # ── Aplicar clasificación a los registros de VPCs ──
    # Construir set de VPCs conectados a hubs (vía TGW o Peering)
    hub_connected_vpcs = set()
    for att in store.tgw_attachments:
        if att["ResourceType"] == "vpc" and att["ResourceId"] not in hub_vpc_ids:
            # Verificar si comparte TGW con un Hub
            for other_att in store.tgw_attachments:
                if (
                    other_att["TgwId"] == att["TgwId"]
                    and other_att["ResourceType"] == "vpc"
                    and other_att["ResourceId"] in hub_vpc_ids
                ):
                    hub_connected_vpcs.add(att["ResourceId"])
                    break

    for pcx in store.vpc_peerings:
        if pcx["Status"] != "active":
            continue
        req_vpc = pcx["RequesterVpcId"]
        acc_vpc = pcx["AccepterVpcId"]
        if req_vpc in hub_vpc_ids:
            hub_connected_vpcs.add(acc_vpc)
        if acc_vpc in hub_vpc_ids:
            hub_connected_vpcs.add(req_vpc)

    # Asignar flags
    for vpc in store.vpcs:
        vpc_id = vpc["VpcId"]
        if vpc_id in hub_vpc_ids:
            vpc["IsHub"] = True
            vpc["HubReasons"] = hub_vpc_ids[vpc_id]
        elif vpc_id in hub_connected_vpcs:
            vpc["IsSpoke"] = True
        # VPCs que no son hub ni spoke conectado quedan sin clasificar (aislados)

    hub_count = sum(1 for v in store.vpcs if v["IsHub"])
    spoke_count = sum(1 for v in store.vpcs if v["IsSpoke"])
    unclass = sum(1 for v in store.vpcs if not v["IsHub"] and not v["IsSpoke"])
    write_status(
        f"  Hubs: {hub_count} | Spokes: {spoke_count} | Sin clasificar: {unclass}", "OK"
    )


# ─────────────────────────────────────────────────────────────────────────────
# PILAR 3 — EVALUACIÓN DE MEJORES PRÁCTICAS
# ─────────────────────────────────────────────────────────────────────────────

def evaluate_best_practices(store: DiscoveryStore) -> None:
    """Ejecuta todas las validaciones de mejores prácticas."""

    write_status("═══════════════════════════════════════════════════════════")
    write_status("  OBJETIVO A — Evaluación de Mejores Prácticas")
    write_status("═══════════════════════════════════════════════════════════")

    bp = store.best_practices
    hub_vpc_ids = {v["VpcId"] for v in store.vpcs if v["IsHub"]}
    spoke_vpc_ids = {v["VpcId"] for v in store.vpcs if v["IsSpoke"]}

    # ── CHECK 1: Conectividad de Spokes — Enrutados hacia TGW/Hub ────────
    write_status("\n[CHECK 1] Validando conectividad de Spokes hacia el Hub/TGW...")

    # 1a: Para cada Spoke, verificar que tenga attachment TGW a un TGW compartido con Hub
    for spoke_vpc in [v for v in store.vpcs if v["IsSpoke"]]:
        vpc_id = spoke_vpc["VpcId"]
        vpc_name = spoke_vpc["VpcName"]

        # Buscar attachments TGW de este spoke
        spoke_tgw_atts = [
            a for a in store.tgw_attachments
            if a["ResourceType"] == "vpc" and a["ResourceId"] == vpc_id
        ]
        # Buscar peerings de este spoke
        spoke_peerings = [
            p for p in store.vpc_peerings
            if p["Status"] == "active" and (p["RequesterVpcId"] == vpc_id or p["AccepterVpcId"] == vpc_id)
        ]

        if not spoke_tgw_atts and not spoke_peerings:
            bp.append({
                "Check": "Spoke-Connectivity",
                "Resource": vpc_name,
                "Status": "WARNING",
                "Detail": "Spoke VPC no tiene TGW attachment ni VPC Peering (aislada)",
                "Recommendation": "Crear TGW attachment o Peering hacia la VPC Hub",
            })
            write_status(f"  ⚠ {vpc_name}: Sin conectividad al Hub", "WARN")
            continue

        # Verificar si tiene ruta 0.0.0.0/0 → TGW en sus route tables
        vpc_rts = [rt for rt in store.route_tables if rt["VpcId"] == vpc_id]
        has_tgw_default_route = any(
            "0.0.0.0/0→tgw-" in rt["Routes"] for rt in vpc_rts
        )
        if spoke_tgw_atts and not has_tgw_default_route:
            bp.append({
                "Check": "Spoke-Connectivity",
                "Resource": vpc_name,
                "Status": "WARNING",
                "Detail": "Spoke tiene TGW attachment pero no tiene ruta 0.0.0.0/0 → TGW",
                "Recommendation": "Agregar ruta por defecto (0.0.0.0/0) apuntando al Transit Gateway",
            })
            write_status(f"  ⚠ {vpc_name}: Sin ruta default hacia TGW", "WARN")
        elif spoke_tgw_atts:
            bp.append({
                "Check": "Spoke-Connectivity",
                "Resource": vpc_name,
                "Status": "PASS",
                "Detail": "Spoke conectado al TGW con ruta default hacia Hub",
                "Recommendation": "N/A",
            })
            write_status(f"  ✓ {vpc_name}: Conectado al TGW con ruta default", "OK")

    # 1b: Detectar VPC Peerings directos entre Spokes (bypassing firewall)
    write_status("\n[CHECK 1b] Detectando Peerings Spoke-to-Spoke (bypass de firewall)...")
    for pcx in store.vpc_peerings:
        if pcx["Status"] != "active":
            continue
        req = pcx["RequesterVpcId"]
        acc = pcx["AccepterVpcId"]
        req_is_spoke = req in spoke_vpc_ids
        acc_is_spoke = acc in spoke_vpc_ids

        if req_is_spoke and acc_is_spoke:
            req_name = _vpc_name(store, req)
            acc_name = _vpc_name(store, acc)
            bp.append({
                "Check": "Spoke-to-Spoke-Peering",
                "Resource": f"{req_name} ↔ {acc_name}",
                "Status": "FAIL",
                "Detail": "Peering directo entre Spokes evade el firewall central del Hub",
                "Recommendation": "Eliminar el peering y enrutar tráfico vía TGW/Hub para inspección",
            })
            write_status(
                f"  ✗ {req_name} ↔ {acc_name}: Peering Spoke-to-Spoke", "ERROR"
            )

    # ── CHECK 2: HA en TGW Attachments (≥2 AZs) ─────────────────────────
    write_status("\n[CHECK 2] Validando Alta Disponibilidad en TGW Attachments (≥2 AZs)...")
    for att in store.tgw_attachments:
        if att["ResourceType"] != "vpc":
            continue
        az_count = att.get("VpcAttachmentAZCount", 0)
        vpc_name = _vpc_name(store, att["ResourceId"])
        resource_label = f"{vpc_name} ({att['TgwAttachmentId']})"

        if az_count >= 2:
            bp.append({
                "Check": "TGW-Attachment-HA",
                "Resource": resource_label,
                "Status": "PASS",
                "Detail": f"Attachment en {az_count} AZs",
                "Recommendation": "N/A",
            })
            write_status(f"  ✓ {resource_label}: {az_count} AZs", "OK")
        elif az_count == 1:
            bp.append({
                "Check": "TGW-Attachment-HA",
                "Resource": resource_label,
                "Status": "FAIL",
                "Detail": f"Attachment en solo {az_count} AZ — sin alta disponibilidad",
                "Recommendation": "Agregar subnets en al menos 2 AZs al TGW attachment",
            })
            write_status(f"  ✗ {resource_label}: Solo {az_count} AZ", "ERROR")
        else:
            bp.append({
                "Check": "TGW-Attachment-HA",
                "Resource": resource_label,
                "Status": "WARNING",
                "Detail": "No se pudo determinar el número de AZs del attachment",
                "Recommendation": "Verificar manualmente la configuración del TGW attachment",
            })

    # ── CHECK 3: HA en Network Firewall y NAT Gateways del Hub ───────────
    write_status("\n[CHECK 3] Validando HA en Network Firewall / NAT Gateways del Hub...")

    for nfw in store.network_firewalls:
        if nfw["VpcId"] not in hub_vpc_ids:
            continue
        ep_count = nfw["EndpointCount"]
        azs = nfw["EndpointAZs"]
        az_list = [az.strip() for az in azs.split(",") if az.strip()] if azs else []

        if len(az_list) >= 2:
            bp.append({
                "Check": "NFW-Hub-HA",
                "Resource": nfw["FirewallName"],
                "Status": "PASS",
                "Detail": f"Network Firewall desplegado en {len(az_list)} AZs: {azs}",
                "Recommendation": "N/A",
            })
            write_status(f"  ✓ {nfw['FirewallName']}: {len(az_list)} AZs", "OK")
        else:
            bp.append({
                "Check": "NFW-Hub-HA",
                "Resource": nfw["FirewallName"],
                "Status": "FAIL",
                "Detail": f"Network Firewall solo en {len(az_list)} AZ(s) — sin HA",
                "Recommendation": "Desplegar endpoint del firewall en al menos 2 AZs",
            })
            write_status(f"  ✗ {nfw['FirewallName']}: Solo {len(az_list)} AZ", "ERROR")

    # NAT Gateways en Hubs
    for hub_vpc in [v for v in store.vpcs if v["IsHub"]]:
        vpc_id = hub_vpc["VpcId"]
        vpc_name = hub_vpc["VpcName"]
        hub_nats = [n for n in store.nat_gateways if n["VpcId"] == vpc_id and n["State"] == "available"]
        if not hub_nats:
            continue  # Hub sin NATs no necesariamente es un problema

        nat_azs = set()
        for nat in hub_nats:
            for sn in store.subnets:
                if sn["SubnetId"] == nat["SubnetId"]:
                    nat_azs.add(sn["AvailabilityZone"])
                    break

        if len(nat_azs) >= 2:
            bp.append({
                "Check": "NAT-Hub-HA",
                "Resource": vpc_name,
                "Status": "PASS",
                "Detail": f"{len(hub_nats)} NAT Gateways en {len(nat_azs)} AZs",
                "Recommendation": "N/A",
            })
            write_status(f"  ✓ Hub {vpc_name}: NATs en {len(nat_azs)} AZs", "OK")
        else:
            bp.append({
                "Check": "NAT-Hub-HA",
                "Resource": vpc_name,
                "Status": "FAIL",
                "Detail": f"NAT Gateways solo en {len(nat_azs)} AZ — punto único de fallo",
                "Recommendation": "Desplegar NAT Gateways en al menos 2 AZs para HA",
            })
            write_status(f"  ✗ Hub {vpc_name}: NATs solo en {len(nat_azs)} AZ", "ERROR")

    # ── CHECK 4: Appliance Mode en TGW Attachment del Hub de Inspección ──
    write_status("\n[CHECK 4] Validando Appliance Mode en TGW Attachments de VPCs Hub/Inspección...")

    for att in store.tgw_attachments:
        if att["ResourceType"] != "vpc":
            continue
        vpc_id = att["ResourceId"]
        if vpc_id not in hub_vpc_ids:
            continue

        vpc_name = _vpc_name(store, vpc_id)
        resource_label = f"{vpc_name} ({att['TgwAttachmentId']})"
        appliance_mode = att.get("ApplianceModeSupport", "unknown")

        # Solo relevante si el Hub tiene Network Firewall
        hub_has_nfw = any(
            nfw["VpcId"] == vpc_id for nfw in store.network_firewalls
        )

        if not hub_has_nfw:
            continue  # No aplica si no hay firewall en este hub

        if appliance_mode == "enable":
            bp.append({
                "Check": "TGW-ApplianceMode",
                "Resource": resource_label,
                "Status": "PASS",
                "Detail": "Appliance Mode habilitado — simetría de tráfico garantizada",
                "Recommendation": "N/A",
            })
            write_status(f"  ✓ {resource_label}: Appliance Mode ON", "OK")
        else:
            bp.append({
                "Check": "TGW-ApplianceMode",
                "Resource": resource_label,
                "Status": "FAIL",
                "Detail": f"Appliance Mode = '{appliance_mode}' — tráfico asimétrico puede bypasear el firewall",
                "Recommendation": (
                    "Habilitar Appliance Mode en el TGW VPC Attachment de la VPC de inspección "
                    "para garantizar que los flujos de tráfico sean simétricos"
                ),
            })
            write_status(f"  ✗ {resource_label}: Appliance Mode OFF", "ERROR")

    # ── CHECK 5: VPN Tunnels Redundancy ──────────────────────────────────
    write_status("\n[CHECK 5] Validando redundancia en conexiones VPN...")
    for vpn in store.vpn_connections:
        if vpn["State"] != "available":
            continue
        statuses = [s.strip() for s in vpn["TunnelStatuses"].split(",") if s.strip()]
        up_count = sum(1 for s in statuses if s == "UP")
        vpn_label = f"{vpn['VpnName']} ({vpn['VpnId']})"

        if up_count >= 2:
            bp.append({
                "Check": "VPN-TunnelRedundancy",
                "Resource": vpn_label,
                "Status": "PASS",
                "Detail": f"{up_count}/{len(statuses)} túneles activos",
                "Recommendation": "N/A",
            })
            write_status(f"  ✓ {vpn_label}: {up_count} túneles UP", "OK")
        elif up_count == 1:
            bp.append({
                "Check": "VPN-TunnelRedundancy",
                "Resource": vpn_label,
                "Status": "WARNING",
                "Detail": f"Solo {up_count}/{len(statuses)} túnel activo — sin redundancia",
                "Recommendation": "Verificar configuración del segundo túnel VPN para HA",
            })
            write_status(f"  ⚠ {vpn_label}: Solo {up_count} túnel UP", "WARN")
        else:
            bp.append({
                "Check": "VPN-TunnelRedundancy",
                "Resource": vpn_label,
                "Status": "FAIL",
                "Detail": f"0/{len(statuses)} túneles activos — VPN sin conectividad",
                "Recommendation": "Investigar y restablecer los túneles VPN inmediatamente",
            })
            write_status(f"  ✗ {vpn_label}: 0 túneles UP", "ERROR")

    # ── CHECK 6: Direct Connect Redundancy ───────────────────────────────
    write_status("\n[CHECK 6] Validando redundancia en Direct Connect...")
    dx_by_location = defaultdict(list)
    for dx in store.dx_connections:
        dx_by_location[dx["Location"]].append(dx)

    for location, connections in dx_by_location.items():
        available_conns = [c for c in connections if c["State"] == "available"]
        if len(available_conns) >= 2:
            bp.append({
                "Check": "DX-Redundancy",
                "Resource": f"Location: {location}",
                "Status": "PASS",
                "Detail": f"{len(available_conns)} conexiones DX activas en esta ubicación",
                "Recommendation": "N/A",
            })
            write_status(f"  ✓ DX {location}: {len(available_conns)} conexiones", "OK")
        elif len(available_conns) == 1:
            bp.append({
                "Check": "DX-Redundancy",
                "Resource": f"Location: {location}",
                "Status": "WARNING",
                "Detail": "Solo 1 conexión DX — sin redundancia a nivel de enlace",
                "Recommendation": (
                    "Agregar una segunda conexión DX (idealmente en otra ubicación "
                    "para resiliencia geográfica)"
                ),
            })
            write_status(f"  ⚠ DX {location}: Solo 1 conexión", "WARN")

    # ── Resumen ──
    pass_count = sum(1 for b in bp if b["Status"] == "PASS")
    fail_count = sum(1 for b in bp if b["Status"] == "FAIL")
    warn_count = sum(1 for b in bp if b["Status"] in ("WARNING", "WARN"))

    write_status("\n═══════════════════════════════════════════════════════════")
    write_status(
        f"  RESUMEN: {pass_count} PASS | {fail_count} FAIL | {warn_count} WARNINGS"
    )
    write_status("═══════════════════════════════════════════════════════════")


def _vpc_name(store: DiscoveryStore, vpc_id: str) -> str:
    """Retorna el nombre legible de un VPC dado su ID."""
    for v in store.vpcs:
        if v["VpcId"] == vpc_id:
            return v["VpcName"]
    return vpc_id


# ─────────────────────────────────────────────────────────────────────────────
# PILAR 4 — GENERACIÓN DE DIAGRAMA MERMAID.JS
# ─────────────────────────────────────────────────────────────────────────────

def generate_mermaid(store: DiscoveryStore) -> str:
    """Genera un diagrama Mermaid.js de la topología Hub & Spoke."""

    write_status("═══════════════════════════════════════════════════════════")
    write_status("  OBJETIVO B — Generación de Diagrama Mermaid.js")
    write_status("═══════════════════════════════════════════════════════════")

    lines = []
    lines.append("graph TD")
    lines.append("")
    lines.append("    %% ═══════════════════════════════════════════════")
    lines.append("    %% AWS Hub & Spoke — Diagrama Auto-Generado")
    lines.append(f"    %% Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    lines.append("    %% ═══════════════════════════════════════════════")
    lines.append("")

    # ── Estilos ──
    lines.append("    %% Estilos")
    lines.append("    classDef hubStyle fill:#1a73e8,stroke:#0d47a1,stroke-width:3px,color:#fff")
    lines.append("    classDef spokeStyle fill:#34a853,stroke:#1b5e20,stroke-width:2px,color:#fff")
    lines.append("    classDef nfwStyle fill:#ea4335,stroke:#b71c1c,stroke-width:2px,color:#fff")
    lines.append("    classDef tgwStyle fill:#ff6d00,stroke:#e65100,stroke-width:3px,color:#fff")
    lines.append("    classDef vpnStyle fill:#fbbc04,stroke:#f57f17,stroke-width:2px,color:#000")
    lines.append("    classDef dnsStyle fill:#9c27b0,stroke:#4a148c,stroke-width:2px,color:#fff")
    lines.append("    classDef natStyle fill:#00bcd4,stroke:#006064,stroke-width:2px,color:#fff")
    lines.append("    classDef onpremStyle fill:#607d8b,stroke:#263238,stroke-width:2px,color:#fff")
    lines.append("    classDef dxStyle fill:#4caf50,stroke:#1b5e20,stroke-width:2px,color:#fff")
    lines.append("    classDef standaloneStyle fill:#78909c,stroke:#37474f,stroke-width:2px,color:#fff")
    lines.append("")

    hub_vpc_ids = {v["VpcId"] for v in store.vpcs if v["IsHub"]}

    # ── Nodo On-Premises (si hay VPN o DX) ──
    has_onprem = (
        len(store.vpn_connections) > 0
        or len(store.dx_connections) > 0
        or len(store.vgws) > 0
    )
    if has_onprem:
        lines.append('    %% On-Premises / Conectividad Externa')
        lines.append('    ONPREM[/"🏢 On-Premises<br/>Datacenter"/]:::onpremStyle')
        lines.append("")

    # ── Transit Gateways ──
    for tgw in store.tgws:
        tgw_id = sanitize_mermaid_id(tgw["TgwId"])
        tgw_label = tgw["TgwName"] if tgw["TgwName"] != tgw["TgwId"] else tgw["TgwId"]
        lines.append(f'    %% Transit Gateway: {tgw_label}')
        lines.append(
            f'    {tgw_id}{{"🔀 TGW: {tgw_label}<br/>'
            f'ASN: {tgw["AmazonSideAsn"]}<br/>'
            f'{tgw["Region"]}"}}:::tgwStyle'
        )
        lines.append("")

    # ── Subgrafos Hub ──
    for hub in [v for v in store.vpcs if v["IsHub"]]:
        hub_id = sanitize_mermaid_id(hub["VpcId"])
        hub_label = hub["VpcName"]
        cidr = hub["CidrBlock"]
        lines.append(f"    %% ── Hub: {hub_label} ──")
        lines.append(f'    subgraph {hub_id}_sub["🔷 HUB: {hub_label}"]')
        lines.append("        direction TB")
        lines.append(
            f'        {hub_id}["📡 {hub_label}<br/>{cidr}<br/>{hub["Region"]}"]:::hubStyle'
        )

        # Network Firewalls
        for nfw in store.network_firewalls:
            if nfw["VpcId"] == hub["VpcId"]:
                nfw_id = sanitize_mermaid_id(nfw["FirewallName"])
                azs_label = nfw["EndpointAZs"] or "N/A"
                lines.append(
                    f'        {nfw_id}["🔥 {nfw["FirewallName"]}<br/>'
                    f'Endpoints: {nfw["EndpointCount"]}<br/>'
                    f'AZs: {azs_label}"]:::nfwStyle'
                )
                lines.append(f"        {hub_id} --- {nfw_id}")

        # Route 53 Resolver Endpoints
        for ep in store.r53_resolver_endpoints:
            if ep["VpcId"] == hub["VpcId"]:
                ep_id = sanitize_mermaid_id(ep["EndpointId"])
                lines.append(
                    f'        {ep_id}["🌐 {ep["EndpointName"]}<br/>'
                    f'R53 Resolver {ep["Direction"]}<br/>'
                    f'IPs: {ep["IpAddressCount"]}"]:::dnsStyle'
                )
                lines.append(f"        {hub_id} --- {ep_id}")

        # NAT Gateways
        hub_nats = [n for n in store.nat_gateways if n["VpcId"] == hub["VpcId"] and n["State"] == "available"]
        if hub_nats:
            if len(hub_nats) <= 4:
                for nat in hub_nats:
                    nat_id = sanitize_mermaid_id(nat["NatGatewayId"])
                    lines.append(
                        f'        {nat_id}["🌊 NAT<br/>{nat["NatGatewayName"]}<br/>'
                        f'{nat["PublicIPs"]}"]:::natStyle'
                    )
                    lines.append(f"        {hub_id} --- {nat_id}")
            else:
                nat_summary_id = f"{hub_id}_nats"
                lines.append(
                    f'        {nat_summary_id}["🌊 {len(hub_nats)} NAT Gateways"]:::natStyle'
                )
                lines.append(f"        {hub_id} --- {nat_summary_id}")

        # VGWs
        for vgw in store.vgws:
            if hub["VpcId"] in vgw["AttachedVpcs"]:
                vgw_id = sanitize_mermaid_id(vgw["VgwId"])
                lines.append(
                    f'        {vgw_id}["🔒 VGW: {vgw["VgwName"]}<br/>'
                    f'ASN: {vgw["AmazonSideAsn"]}"]:::vpnStyle'
                )
                lines.append(f"        {hub_id} --- {vgw_id}")

        lines.append("    end")
        lines.append("")

    # ── Subgrafos Spoke ──
    for spoke in [v for v in store.vpcs if v["IsSpoke"]]:
        spoke_id = sanitize_mermaid_id(spoke["VpcId"])
        spoke_label = spoke["VpcName"]
        cidr = spoke["CidrBlock"]
        lines.append(f"    %% ── Spoke: {spoke_label} ──")
        lines.append(f'    subgraph {spoke_id}_sub["🟢 SPOKE: {spoke_label}"]')
        lines.append(
            f'        {spoke_id}["🖥️ {spoke_label}<br/>{cidr}<br/>{spoke["Region"]}"]:::spokeStyle'
        )

        # Subnets del Spoke (limitar a 5 por legibilidad)
        spoke_subnets = [
            sn for sn in store.subnets if sn["VpcId"] == spoke["VpcId"]
        ]
        if 0 < len(spoke_subnets) <= 5:
            for sn in spoke_subnets:
                sn_id = sanitize_mermaid_id(f"{spoke['VpcId']}_{sn['SubnetId']}")
                lines.append(
                    f'        {sn_id}["📂 {sn["SubnetName"]}<br/>{sn["CidrBlock"]}"]'
                )
                lines.append(f"        {spoke_id} --- {sn_id}")
        elif len(spoke_subnets) > 5:
            sn_summary = f"{spoke_id}_sn"
            lines.append(f'        {sn_summary}["📂 {len(spoke_subnets)} subnets"]')
            lines.append(f"        {spoke_id} --- {sn_summary}")

        lines.append("    end")
        lines.append("")

    # ── Subgrafos VPCs sin clasificar (ni Hub ni Spoke) ──
    unclassified_vpcs = [v for v in store.vpcs if not v["IsHub"] and not v["IsSpoke"]]
    if unclassified_vpcs:
        lines.append("    %% ═══ VPCs sin clasificar (Standalone) ═══")
        for vpc in unclassified_vpcs:
            vpc_id = sanitize_mermaid_id(vpc["VpcId"])
            vpc_label = vpc["VpcName"]
            cidr = vpc["CidrBlock"]
            default_tag = " [Default]" if vpc.get("IsDefault") else ""
            lines.append(f"    %% ── Standalone: {vpc_label} ──")
            lines.append(f'    subgraph {vpc_id}_sub["⬜ VPC: {vpc_label}{default_tag}"]')
            lines.append("        direction TB")
            lines.append(
                f'        {vpc_id}["🖥️ {vpc_label}<br/>{cidr}<br/>'
                f'{vpc["Region"]}{default_tag}"]:::standaloneStyle'
            )

            # Subnets
            vpc_subnets = [sn for sn in store.subnets if sn["VpcId"] == vpc["VpcId"]]
            if 0 < len(vpc_subnets) <= 6:
                for sn in vpc_subnets:
                    sn_id = sanitize_mermaid_id(f"{vpc['VpcId']}_{sn['SubnetId']}")
                    pub_label = " 🌐" if sn.get("MapPublicIp") else ""
                    lines.append(
                        f'        {sn_id}["📂 {sn["SubnetName"]}<br/>'
                        f'{sn["CidrBlock"]}<br/>{sn["AvailabilityZone"]}{pub_label}"]'
                    )
                    lines.append(f"        {vpc_id} --- {sn_id}")
            elif len(vpc_subnets) > 6:
                sn_summary = f"{vpc_id}_sn"
                # Agrupar por AZ
                azs = set(sn["AvailabilityZone"] for sn in vpc_subnets)
                lines.append(
                    f'        {sn_summary}["📂 {len(vpc_subnets)} subnets<br/>'
                    f'{len(azs)} AZs"]'
                )
                lines.append(f"        {vpc_id} --- {sn_summary}")

            # NAT Gateways
            vpc_nats = [
                n for n in store.nat_gateways
                if n["VpcId"] == vpc["VpcId"] and n["State"] == "available"
            ]
            for nat in vpc_nats:
                nat_id = sanitize_mermaid_id(nat["NatGatewayId"])
                lines.append(
                    f'        {nat_id}["🌊 NAT<br/>{nat["NatGatewayName"]}<br/>'
                    f'{nat["PublicIPs"]}"]:::natStyle'
                )
                lines.append(f"        {vpc_id} --- {nat_id}")

            # VGWs
            for vgw in store.vgws:
                if vpc["VpcId"] in vgw["AttachedVpcs"]:
                    vgw_id = sanitize_mermaid_id(vgw["VgwId"])
                    lines.append(
                        f'        {vgw_id}["🔒 VGW: {vgw["VgwName"]}<br/>'
                        f'ASN: {vgw["AmazonSideAsn"]}"]:::vpnStyle'
                    )
                    lines.append(f"        {vpc_id} --- {vgw_id}")

            # R53 Resolver Endpoints
            for ep in store.r53_resolver_endpoints:
                if ep["VpcId"] == vpc["VpcId"]:
                    ep_id = sanitize_mermaid_id(ep["EndpointId"])
                    lines.append(
                        f'        {ep_id}["🌐 {ep["EndpointName"]}<br/>'
                        f'R53 {ep["Direction"]}"]:::dnsStyle'
                    )
                    lines.append(f"        {vpc_id} --- {ep_id}")

            lines.append("    end")
            lines.append("")

    # ── Conexiones TGW ↔ VPC ──
    lines.append("    %% ═══ TGW Attachments ═══")
    for att in store.tgw_attachments:
        if att["ResourceType"] == "vpc":
            tgw_mid = sanitize_mermaid_id(att["TgwId"])
            vpc_mid = sanitize_mermaid_id(att["ResourceId"])
            vpc_id = att["ResourceId"]
            az_count = att.get("VpcAttachmentAZCount", "?")
            appliance = att.get("ApplianceModeSupport", "?")

            if vpc_id in hub_vpc_ids:
                label = f"Hub Att<br/>{az_count} AZs"
                if appliance == "enable":
                    label += "<br/>Appliance ✓"
                lines.append(f"    {tgw_mid} <==>|{label}| {vpc_mid}")
            else:
                label = f"Spoke Att<br/>{az_count} AZs"
                lines.append(f"    {tgw_mid} <-->|{label}| {vpc_mid}")

    lines.append("")

    # ── Conexiones VPN / DX → On-Premises ──
    if has_onprem:
        lines.append("    %% ═══ On-Premises Connectivity ═══")
        for vpn in store.vpn_connections:
            if vpn["TgwId"]:
                tgw_mid = sanitize_mermaid_id(vpn["TgwId"])
                lines.append(f'    ONPREM -.->|"VPN: {vpn["VpnName"]}"| {tgw_mid}')
            elif vpn["VgwId"]:
                vgw_mid = sanitize_mermaid_id(vpn["VgwId"])
                lines.append(f'    ONPREM -.->|"VPN: {vpn["VpnName"]}"| {vgw_mid}')

        for dx in store.dx_connections:
            if dx["State"] == "available":
                # DX normalmente se conecta vía Direct Connect Gateway → TGW
                # Buscar un TGW attachment tipo direct-connect-gateway
                dx_tgw = None
                for att in store.tgw_attachments:
                    if att["ResourceType"] == "direct-connect-gateway":
                        dx_tgw = att["TgwId"]
                        break
                if dx_tgw:
                    tgw_mid = sanitize_mermaid_id(dx_tgw)
                    lines.append(
                        f'    ONPREM ==>|"DX: {dx["ConnectionName"]}<br/>'
                        f'{dx["Bandwidth"]}"| {tgw_mid}'
                    )
                else:
                    lines.append(
                        f'    ONPREM ==>|"DX: {dx["ConnectionName"]}<br/>'
                        f'{dx["Bandwidth"]}"| DX_GW'
                    )
                    # Nodo DX Gateway genérico si no hay TGW
                    if 'DX_GW' not in "\n".join(lines):
                        lines.insert(
                            lines.index("    %% ═══ On-Premises Connectivity ═══"),
                            '    DX_GW{{"🔌 Direct Connect Gateway"}}:::dxStyle'
                        )

        lines.append("")

    # ── Peerings VPC ──
    lines.append("    %% ═══ VPC Peerings ═══")
    processed_peerings = set()
    for pcx in store.vpc_peerings:
        if pcx["Status"] != "active":
            continue
        pair = tuple(sorted([pcx["RequesterVpcId"], pcx["AccepterVpcId"]]))
        if pair in processed_peerings:
            continue
        processed_peerings.add(pair)

        req_mid = sanitize_mermaid_id(pcx["RequesterVpcId"])
        acc_mid = sanitize_mermaid_id(pcx["AccepterVpcId"])
        req_name = _vpc_name(store, pcx["RequesterVpcId"])
        acc_name = _vpc_name(store, pcx["AccepterVpcId"])

        req_is_hub = pcx["RequesterVpcId"] in hub_vpc_ids
        acc_is_hub = pcx["AccepterVpcId"] in hub_vpc_ids

        if req_is_hub or acc_is_hub:
            lines.append(f"    {req_mid} <-->|Peering| {acc_mid}")
        else:
            lines.append(f"    {req_mid} -.->|S2S Peering| {acc_mid}")

    lines.append("")
    mermaid_content = "\n".join(lines)
    write_status("Diagrama Mermaid generado correctamente", "OK")
    return mermaid_content


# ─────────────────────────────────────────────────────────────────────────────
# PILAR 5 — EXPORTACIÓN (CSVs + Mermaid + Markdown Summary)
# ─────────────────────────────────────────────────────────────────────────────

def export_csv(data: list[dict], filepath: str, description: str) -> None:
    """Exporta una lista de diccionarios a un archivo CSV (fallback si no hay openpyxl)."""
    if not data:
        write_status(f"  {description}: Sin datos, omitiendo", "WARN")
        return
    exclude_keys = {"HubReasons"}
    fieldnames = [k for k in data[0].keys() if k not in exclude_keys]
    try:
        with open(filepath, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(data)
        write_status(f"  {description}: {filepath}", "OK")
    except Exception as e:
        write_status(f"  Error exportando {filepath}: {e}", "ERROR")


# ── Estilos Excel reutilizables ──────────────────────────────────────────

_HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
_HEADER_FILL_BLUE = PatternFill("solid", fgColor="1A73E8")
_HEADER_FILL_GREEN = PatternFill("solid", fgColor="34A853")
_HEADER_FILL_RED = PatternFill("solid", fgColor="EA4335")
_HEADER_FILL_ORANGE = PatternFill("solid", fgColor="FF6D00")
_HEADER_FILL_GRAY = PatternFill("solid", fgColor="455A64")
_HEADER_FILL_PURPLE = PatternFill("solid", fgColor="7B1FA2")
_HEADER_FILL_TEAL = PatternFill("solid", fgColor="00838F")
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
_CELL_FONT = Font(name="Arial", size=10)
_CELL_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
_THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)
_PASS_FILL = PatternFill("solid", fgColor="E8F5E9")
_FAIL_FILL = PatternFill("solid", fgColor="FFEBEE")
_WARN_FILL = PatternFill("solid", fgColor="FFF8E1")
_PASS_FONT = Font(name="Arial", size=10, bold=True, color="1B5E20")
_FAIL_FONT = Font(name="Arial", size=10, bold=True, color="B71C1C")
_WARN_FONT = Font(name="Arial", size=10, bold=True, color="E65100")


def _add_sheet_from_data(
    wb: "Workbook",
    sheet_name: str,
    data: list[dict],
    header_fill: PatternFill = _HEADER_FILL_BLUE,
    exclude_keys: set = None,
    is_best_practices: bool = False,
) -> None:
    """Agrega una hoja al workbook con datos, encabezados formateados y auto-width."""
    if not data:
        return

    exclude_keys = exclude_keys or set()
    fieldnames = [k for k in data[0].keys() if k not in exclude_keys]

    ws = wb.create_sheet(title=sheet_name[:31])  # Excel limita a 31 chars

    # Encabezados
    for col_idx, header in enumerate(fieldnames, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = header_fill
        cell.alignment = _HEADER_ALIGNMENT
        cell.border = _THIN_BORDER

    # Datos
    for row_idx, record in enumerate(data, 2):
        for col_idx, key in enumerate(fieldnames, 1):
            value = record.get(key, "")
            # Convertir listas a string
            if isinstance(value, (list, tuple)):
                value = ", ".join(str(v) for v in value)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = _CELL_FONT
            cell.alignment = _CELL_ALIGNMENT
            cell.border = _THIN_BORDER

            # Colorear filas de Best Practices según Status
            if is_best_practices and key == "Status":
                if value == "PASS":
                    cell.font = _PASS_FONT
                    cell.fill = _PASS_FILL
                elif value == "FAIL":
                    cell.font = _FAIL_FONT
                    cell.fill = _FAIL_FILL
                elif value in ("WARNING", "WARN"):
                    cell.font = _WARN_FONT
                    cell.fill = _WARN_FILL

    # Auto-fit column widths
    for col_idx, key in enumerate(fieldnames, 1):
        max_len = len(str(key))
        for row_idx in range(2, min(len(data) + 2, 102)):  # Muestrear hasta 100 filas
            val = str(data[row_idx - 2].get(key, ""))
            max_len = max(max_len, min(len(val), 60))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 3

    # Filtros automáticos
    ws.auto_filter.ref = ws.dimensions

    # Congelar encabezado
    ws.freeze_panes = "A2"


def _add_summary_sheet(wb: "Workbook", store: "DiscoveryStore") -> None:
    """Agrega la hoja de resumen ejecutivo como primera pestaña."""
    ws = wb.create_sheet(title="Resumen", index=0)

    # Título
    ws.merge_cells("A1:D1")
    title_cell = ws["A1"]
    title_cell.value = "AWS Hub & Spoke — Resumen Arquitectónico"
    title_cell.font = Font(name="Arial", bold=True, size=14, color="1A73E8")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Subtítulo
    ws.merge_cells("A2:D2")
    sub_cell = ws["A2"]
    sub_cell.value = f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M UTC')} — discover_hub_spoke_aws.py v{SCRIPT_VERSION}"
    sub_cell.font = Font(name="Arial", size=9, italic=True, color="666666")
    sub_cell.alignment = Alignment(horizontal="center")

    row = 4
    # Inventario general
    ws.cell(row=row, column=1, value="INVENTARIO GENERAL").font = Font(name="Arial", bold=True, size=11, color="1A73E8")
    row += 1

    summary = store.summary()
    accounts = list(set(v["AccountId"] for v in store.vpcs)) if store.vpcs else []
    regions = list(set(v["Region"] for v in store.vpcs)) if store.vpcs else []

    summary_data = [
        ("Cuentas analizadas", len(accounts)),
        ("Regiones", ", ".join(regions) if regions else "N/A"),
    ]
    summary_data.extend(summary.items())

    for metric, value in summary_data:
        ws.cell(row=row, column=1, value=metric).font = Font(name="Arial", size=10)
        ws.cell(row=row, column=1).border = _THIN_BORDER
        val_cell = ws.cell(row=row, column=2, value=value)
        val_cell.font = Font(name="Arial", size=10, bold=True)
        val_cell.border = _THIN_BORDER
        row += 1

    row += 1

    # Topología
    hub_count = sum(1 for v in store.vpcs if v["IsHub"])
    spoke_count = sum(1 for v in store.vpcs if v["IsSpoke"])
    unclass_count = sum(1 for v in store.vpcs if not v["IsHub"] and not v["IsSpoke"])

    ws.cell(row=row, column=1, value="TOPOLOGÍA").font = Font(name="Arial", bold=True, size=11, color="1A73E8")
    row += 1
    for label, count, color in [
        ("Hubs", hub_count, "1A73E8"),
        ("Spokes", spoke_count, "34A853"),
        ("Sin clasificar", unclass_count, "78909C"),
    ]:
        ws.cell(row=row, column=1, value=label).font = Font(name="Arial", size=10)
        ws.cell(row=row, column=1).border = _THIN_BORDER
        c = ws.cell(row=row, column=2, value=count)
        c.font = Font(name="Arial", size=10, bold=True, color=color)
        c.border = _THIN_BORDER
        row += 1

    row += 1

    # Best Practices resumen
    bp = store.best_practices
    pass_c = sum(1 for b in bp if b["Status"] == "PASS")
    fail_c = sum(1 for b in bp if b["Status"] == "FAIL")
    warn_c = sum(1 for b in bp if b["Status"] in ("WARNING", "WARN"))

    ws.cell(row=row, column=1, value="AUDITORÍA DE MEJORES PRÁCTICAS").font = Font(name="Arial", bold=True, size=11, color="1A73E8")
    row += 1
    for label, count, fill, font in [
        ("PASS", pass_c, _PASS_FILL, _PASS_FONT),
        ("FAIL", fail_c, _FAIL_FILL, _FAIL_FONT),
        ("WARNING", warn_c, _WARN_FILL, _WARN_FONT),
    ]:
        ws.cell(row=row, column=1, value=label).font = font
        ws.cell(row=row, column=1).fill = fill
        ws.cell(row=row, column=1).border = _THIN_BORDER
        c = ws.cell(row=row, column=2, value=count)
        c.font = Font(name="Arial", size=10, bold=True)
        c.border = _THIN_BORDER
        row += 1

    # Hubs detalle
    if hub_count > 0:
        row += 1
        ws.cell(row=row, column=1, value="DETALLE DE HUBS").font = Font(name="Arial", bold=True, size=11, color="1A73E8")
        row += 1
        for hub in [v for v in store.vpcs if v["IsHub"]]:
            ws.cell(row=row, column=1, value=hub["VpcName"]).font = Font(name="Arial", size=10, bold=True)
            ws.cell(row=row, column=2, value=hub["VpcId"]).font = Font(name="Arial", size=9, color="666666")
            ws.cell(row=row, column=3, value=hub["CidrBlock"]).font = _CELL_FONT
            ws.cell(row=row, column=4, value=hub["Region"]).font = _CELL_FONT
            row += 1
            reasons = hub.get("HubReasons", [])
            if reasons:
                ws.cell(row=row, column=2, value="Razones: " + "; ".join(reasons)).font = Font(name="Arial", size=9, italic=True, color="555555")
                row += 1

    # Column widths
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 18


def _add_mermaid_sheet(wb: "Workbook", mermaid_content: str) -> None:
    """Agrega una hoja con el código Mermaid.js para copiar."""
    ws = wb.create_sheet(title="Diagrama Mermaid")

    ws.merge_cells("A1:C1")
    ws["A1"].value = "Diagrama Mermaid.js — Copiar y pegar en https://mermaid.live"
    ws["A1"].font = Font(name="Arial", bold=True, size=11, color="1A73E8")

    ws["A2"].value = "Seleccionar toda la columna A desde la fila 4 para copiar el código."
    ws["A2"].font = Font(name="Arial", size=9, italic=True, color="666666")

    for row_idx, line in enumerate(mermaid_content.split("\n"), 4):
        cell = ws.cell(row=row_idx, column=1, value=line)
        cell.font = Font(name="Consolas", size=9)

    ws.column_dimensions["A"].width = 120


def export_all(store: DiscoveryStore, mermaid_content: str, output_path: str) -> dict:
    """Exporta todos los artefactos: libro Excel consolidado + Mermaid .mmd + Markdown."""

    write_status("\n═══════════════════════════════════════════════════════════")
    write_status("  Exportando resultados...")
    write_status("═══════════════════════════════════════════════════════════")

    os.makedirs(output_path, exist_ok=True)
    files = {}

    # ── Libro Excel consolidado ──
    if HAS_OPENPYXL:
        xlsx_path = os.path.join(output_path, "hub-spoke-discovery.xlsx")
        wb = Workbook()
        # Eliminar la hoja por defecto
        wb.remove(wb.active)

        # 1. Resumen ejecutivo
        _add_summary_sheet(wb, store)

        # 2. Best Practices (con formato condicional por Status)
        _add_sheet_from_data(
            wb, "Best Practices", store.best_practices,
            header_fill=_HEADER_FILL_RED, is_best_practices=True,
        )

        # 3-15. Hojas de inventario
        sheets_config = [
            ("VPCs", store.vpcs, _HEADER_FILL_BLUE, {"HubReasons"}),
            ("Subnets", store.subnets, _HEADER_FILL_BLUE, set()),
            ("Route Tables", store.route_tables, _HEADER_FILL_BLUE, set()),
            ("Transit Gateways", store.tgws, _HEADER_FILL_ORANGE, set()),
            ("TGW Attachments", store.tgw_attachments, _HEADER_FILL_ORANGE, set()),
            ("VPC Peerings", store.vpc_peerings, _HEADER_FILL_GREEN, set()),
            ("VGWs", store.vgws, _HEADER_FILL_PURPLE, set()),
            ("Customer Gateways", store.cgws, _HEADER_FILL_PURPLE, set()),
            ("VPN Connections", store.vpn_connections, _HEADER_FILL_PURPLE, set()),
            ("Direct Connect", store.dx_connections, _HEADER_FILL_TEAL, set()),
            ("Network Firewalls", store.network_firewalls, _HEADER_FILL_RED, set()),
            ("R53 Resolver", store.r53_resolver_endpoints, _HEADER_FILL_TEAL, set()),
            ("NAT Gateways", store.nat_gateways, _HEADER_FILL_TEAL, set()),
        ]

        for sheet_name, data, fill, exclude in sheets_config:
            if data:
                _add_sheet_from_data(wb, sheet_name, data, header_fill=fill, exclude_keys=exclude)
                write_status(f"  Hoja '{sheet_name}': {len(data)} registros", "OK")
            else:
                write_status(f"  Hoja '{sheet_name}': Sin datos, omitiendo", "WARN")

        # Hoja de Mermaid
        _add_mermaid_sheet(wb, mermaid_content)

        wb.save(xlsx_path)
        write_status(f"  📊 Libro Excel: {xlsx_path}", "OK")
        files["xlsx"] = xlsx_path

    else:
        write_status("  openpyxl no disponible — exportando CSVs individuales", "WARN")
        write_status("  Para Excel: pip install openpyxl", "WARN")
        export_csv(store.vpcs, os.path.join(output_path, "inventory-vpcs.csv"), "VPCs")
        export_csv(store.subnets, os.path.join(output_path, "inventory-subnets.csv"), "Subnets")
        export_csv(store.route_tables, os.path.join(output_path, "inventory-route-tables.csv"), "Route Tables")
        export_csv(store.tgws, os.path.join(output_path, "inventory-tgws.csv"), "TGWs")
        export_csv(store.tgw_attachments, os.path.join(output_path, "inventory-tgw-attachments.csv"), "TGW Attachments")
        export_csv(store.vpc_peerings, os.path.join(output_path, "inventory-vpc-peerings.csv"), "VPC Peerings")
        export_csv(store.vgws, os.path.join(output_path, "inventory-vgws.csv"), "VGWs")
        export_csv(store.cgws, os.path.join(output_path, "inventory-cgws.csv"), "Customer GWs")
        export_csv(store.vpn_connections, os.path.join(output_path, "inventory-vpn-connections.csv"), "VPN Connections")
        export_csv(store.dx_connections, os.path.join(output_path, "inventory-dx-connections.csv"), "Direct Connect")
        export_csv(store.network_firewalls, os.path.join(output_path, "inventory-network-firewalls.csv"), "Network Firewalls")
        export_csv(store.r53_resolver_endpoints, os.path.join(output_path, "inventory-r53-resolver.csv"), "R53 Resolver")
        export_csv(store.nat_gateways, os.path.join(output_path, "inventory-nat-gateways.csv"), "NAT Gateways")
        bp_path = os.path.join(output_path, "best-practices-report.csv")
        export_csv(store.best_practices, bp_path, "Best Practices")
        files["best_practices_csv"] = bp_path

    # ── Mermaid .mmd (siempre) ──
    mermaid_path = os.path.join(output_path, "hub-spoke-topology.mmd")
    with open(mermaid_path, "w", encoding="utf-8") as f:
        f.write(mermaid_content)
    write_status(f"  🗺️  Diagrama Mermaid: {mermaid_path}", "OK")
    files["mermaid"] = mermaid_path

    # ── Markdown Summary (siempre) ──
    md_path = os.path.join(output_path, "architecture-summary.md")
    md = _generate_markdown_summary(store, mermaid_content)
    with open(md_path, "w", encoding="utf-8") as f:
        f.write(md)
    write_status(f"  📝 Resumen Markdown: {md_path}", "OK")
    files["summary_md"] = md_path

    return files


def _generate_markdown_summary(store: DiscoveryStore, mermaid_content: str) -> str:
    """Genera el resumen arquitectónico completo en Markdown."""

    lines = []
    now = datetime.now().strftime("%Y-%m-%d %H:%M UTC")

    lines.append("# AWS Hub & Spoke — Resumen Arquitectónico")
    lines.append("")
    lines.append(f"> Documento auto-generado por `discover_hub_spoke_aws.py` el {now}.")
    lines.append("> Este resumen describe la topología de red AWS descubierta y los hallazgos de auditoría.")
    lines.append("")

    # ── 1. Inventario General ──
    lines.append("## 1. Inventario General")
    lines.append("")
    summary = store.summary()
    lines.append("| Métrica | Valor |")
    lines.append("|---|---|")
    accounts = list(set(v["AccountId"] for v in store.vpcs))
    regions = list(set(v["Region"] for v in store.vpcs))
    lines.append(f"| Cuentas analizadas | {len(accounts)} |")
    lines.append(f"| Regiones | {', '.join(regions)} |")
    for metric, value in summary.items():
        lines.append(f"| {metric} | {value} |")
    lines.append("")

    # ── 2. Topología ──
    lines.append("## 2. Topología Descubierta")
    lines.append("")
    hub_vpcs = [v for v in store.vpcs if v["IsHub"]]
    spoke_vpcs = [v for v in store.vpcs if v["IsSpoke"]]

    if not hub_vpcs:
        lines.append("No se identificaron VPCs Hub. Verificar que las cuentas con TGW/NFW estén incluidas.")
    else:
        lines.append(f"La arquitectura presenta **{len(hub_vpcs)} Hub(s)** y **{len(spoke_vpcs)} Spoke(s)**.")
        lines.append("")

        for hub in hub_vpcs:
            lines.append(f"### Hub: {hub['VpcName']}")
            lines.append("")
            lines.append("| Atributo | Valor |")
            lines.append("|---|---|")
            lines.append(f"| Cuenta | {hub['AccountLabel']} ({hub['AccountId']}) |")
            lines.append(f"| Región | {hub['Region']} |")
            lines.append(f"| VPC ID | `{hub['VpcId']}` |")
            lines.append(f"| CIDR | `{hub['CidrBlock']}` |")
            reasons = hub.get("HubReasons", [])
            if reasons:
                lines.append(f"| Razones de clasificación | {'; '.join(reasons)} |")
            lines.append("")

            # NFWs
            hub_nfws = [n for n in store.network_firewalls if n["VpcId"] == hub["VpcId"]]
            if hub_nfws:
                lines.append("**Network Firewalls:**")
                lines.append("")
                lines.append("| Firewall | AZs | Endpoints | Policy ARN |")
                lines.append("|---|---|---|---|")
                for nfw in hub_nfws:
                    lines.append(
                        f"| {nfw['FirewallName']} | {nfw['EndpointAZs']} "
                        f"| {nfw['EndpointCount']} | `{nfw['PolicyArn']}` |"
                    )
                lines.append("")

            # R53 Resolver
            hub_resolvers = [e for e in store.r53_resolver_endpoints if e["VpcId"] == hub["VpcId"]]
            if hub_resolvers:
                lines.append("**Route 53 Resolver Endpoints:**")
                lines.append("")
                for ep in hub_resolvers:
                    lines.append(f"- **{ep['Direction']}**: {ep['EndpointName']} ({ep['IpAddressCount']} IPs)")
                lines.append("")

    # ── 3. Spokes ──
    lines.append("## 3. VPCs Spoke")
    lines.append("")
    if not spoke_vpcs:
        lines.append("No se identificaron VPCs Spoke.")
    else:
        lines.append("| Spoke | Cuenta | Región | CIDR | Subnets |")
        lines.append("|---|---|---|---|---|")
        for spoke in spoke_vpcs:
            sn_count = sum(1 for sn in store.subnets if sn["VpcId"] == spoke["VpcId"])
            lines.append(
                f"| {spoke['VpcName']} | {spoke['AccountId']} "
                f"| {spoke['Region']} | `{spoke['CidrBlock']}` | {sn_count} |"
            )
        lines.append("")

    # ── 4. Conectividad On-Premises ──
    lines.append("## 4. Conectividad On-Premises")
    lines.append("")
    if store.vpn_connections:
        lines.append("### VPN Connections")
        lines.append("")
        lines.append("| VPN | Estado | Tipo | TGW/VGW | CGW | Túneles |")
        lines.append("|---|---|---|---|---|---|")
        for vpn in store.vpn_connections:
            target = vpn["TgwId"] or vpn["VgwId"]
            lines.append(
                f"| {vpn['VpnName']} | {vpn['State']} | {vpn['Type']} "
                f"| `{target}` | `{vpn['CgwId']}` | {vpn['TunnelStatuses']} |"
            )
        lines.append("")

    if store.dx_connections:
        lines.append("### Direct Connect")
        lines.append("")
        lines.append("| Conexión | Estado | Bandwidth | Ubicación | Partner |")
        lines.append("|---|---|---|---|---|")
        for dx in store.dx_connections:
            lines.append(
                f"| {dx['ConnectionName']} | {dx['State']} | {dx['Bandwidth']} "
                f"| {dx['Location']} | {dx['PartnerName']} |"
            )
        lines.append("")

    # ── 5. Transit Gateways ──
    if store.tgws:
        lines.append("## 5. Transit Gateways")
        lines.append("")
        lines.append("| TGW | Región | ASN | Estado | Attachments |")
        lines.append("|---|---|---|---|---|")
        for tgw in store.tgws:
            att_count = sum(
                1 for a in store.tgw_attachments if a["TgwId"] == tgw["TgwId"]
            )
            lines.append(
                f"| {tgw['TgwName']} | {tgw['Region']} | {tgw['AmazonSideAsn']} "
                f"| {tgw['State']} | {att_count} |"
            )
        lines.append("")

    # ── 6. Auditoría Best Practices ──
    lines.append("## 6. Auditoría de Mejores Prácticas")
    lines.append("")
    bp = store.best_practices
    pass_c = sum(1 for b in bp if b["Status"] == "PASS")
    fail_c = sum(1 for b in bp if b["Status"] == "FAIL")
    warn_c = sum(1 for b in bp if b["Status"] in ("WARNING", "WARN"))
    lines.append("| Resultado | Cantidad |")
    lines.append("|---|---|")
    lines.append(f"| PASS | {pass_c} |")
    lines.append(f"| FAIL | {fail_c} |")
    lines.append(f"| WARNING | {warn_c} |")
    lines.append("")

    # Agrupar por Check
    checks = defaultdict(list)
    for b in bp:
        checks[b["Check"]].append(b)

    for check_name, items in checks.items():
        fails = [i for i in items if i["Status"] == "FAIL"]
        warns = [i for i in items if i["Status"] in ("WARNING", "WARN")]
        passes = [i for i in items if i["Status"] == "PASS"]
        emoji = "🔴" if fails else ("🟡" if warns else "🟢")

        lines.append(f"### {emoji} {check_name}")
        lines.append("")
        lines.append(f"{len(passes)} pass, {len(fails)} fail, {len(warns)} warning.")
        lines.append("")

        if fails:
            lines.append("**Hallazgos que requieren acción:**")
            lines.append("")
            lines.append("| Recurso | Detalle | Recomendación |")
            lines.append("|---|---|---|")
            for item in fails:
                lines.append(f"| {item['Resource']} | {item['Detail']} | {item['Recommendation']} |")
            lines.append("")

        if warns:
            lines.append("**Advertencias:**")
            lines.append("")
            lines.append("| Recurso | Detalle | Recomendación |")
            lines.append("|---|---|---|")
            for item in warns:
                lines.append(f"| {item['Resource']} | {item['Detail']} | {item['Recommendation']} |")
            lines.append("")

    # ── 7. Diagrama ──
    lines.append("## 7. Diagrama de Topología")
    lines.append("")
    lines.append("El siguiente diagrama fue generado automáticamente. "
                 "Para renderizarlo, pegar en [mermaid.live](https://mermaid.live).")
    lines.append("")
    lines.append("```mermaid")
    lines.append(mermaid_content)
    lines.append("```")
    lines.append("")

    # ── 8. Recomendaciones ──
    lines.append("## 8. Recomendaciones Generales")
    lines.append("")
    recommendations = []

    if not store.network_firewalls:
        recommendations.append(
            "**Desplegar AWS Network Firewall en el Hub.** No se detectaron Network Firewalls. "
            "Sin inspección centralizada, el tráfico inter-spoke y hacia internet no se filtra. "
            "Considerar desplegar AWS Network Firewall en la VPC Hub con endpoints en múltiples AZs."
        )

    if not store.r53_resolver_endpoints:
        recommendations.append(
            "**Implementar Route 53 Resolver Endpoints.** No se detectaron endpoints de resolución DNS. "
            "Para resolución DNS consistente entre on-premises y AWS, desplegar Inbound y Outbound "
            "endpoints en la VPC Hub."
        )

    s2s_peerings = [
        b for b in bp if b["Check"] == "Spoke-to-Spoke-Peering" and b["Status"] == "FAIL"
    ]
    if s2s_peerings:
        recommendations.append(
            f"**Eliminar peerings Spoke-to-Spoke.** Se detectaron {len(s2s_peerings)} peerings directos "
            "entre VPCs Spoke que evaden la inspección centralizada. Enrutar todo el tráfico inter-spoke "
            "a través del Transit Gateway y el Network Firewall del Hub."
        )

    appliance_fails = [
        b for b in bp if b["Check"] == "TGW-ApplianceMode" and b["Status"] == "FAIL"
    ]
    if appliance_fails:
        recommendations.append(
            f"**Habilitar Appliance Mode.** {len(appliance_fails)} TGW attachment(s) de la VPC de "
            "inspección no tienen Appliance Mode habilitado. Sin esto, el tráfico puede ser asimétrico "
            "y bypasear el firewall."
        )

    if recommendations:
        for i, rec in enumerate(recommendations, 1):
            lines.append(f"{i}. {rec}")
            lines.append("")
    else:
        lines.append("No se identificaron recomendaciones críticas.")
        lines.append("")

    # ── Footer ──
    lines.append("---")
    lines.append("")
    lines.append("*Generado por [discover_hub_spoke_aws.py](discover_hub_spoke_aws.py) — "
                 "Para actualizar, volver a ejecutar el script.*")

    return "\n".join(lines)


# ─────────────────────────────────────────────────────────────────────────────
# MAIN — Punto de entrada
# ─────────────────────────────────────────────────────────────────────────────

def parse_args():
    parser = argparse.ArgumentParser(
        description=(
            "AWS Hub & Spoke Discovery — Descubrimiento, clasificación "
            "y auditoría de arquitectura de red AWS."
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Ejemplos:
  # Cuenta por defecto, regiones específicas
  python3 discover_hub_spoke_aws.py --regions us-east-1,us-west-2

  # Múltiples perfiles de AWS CLI
  python3 discover_hub_spoke_aws.py --profiles prod,dev --regions us-east-1

  # AWS Organizations con rol cross-account
  python3 discover_hub_spoke_aws.py --org --role-arn arn:aws:iam::{account_id}:role/OrgReadOnly

  # Lista explícita de cuentas
  python3 discover_hub_spoke_aws.py --account-ids 111111111111,222222222222 \\
                                    --role-arn arn:aws:iam::{account_id}:role/NetAudit
        """
    )
    parser.add_argument(
        "--profiles",
        help="Perfiles de AWS CLI separados por coma (ej: prod,dev,staging)",
    )
    parser.add_argument(
        "--org",
        action="store_true",
        help="Usar AWS Organizations para listar todas las cuentas activas",
    )
    parser.add_argument(
        "--account-ids",
        help="IDs de cuenta AWS separados por coma (requiere --role-arn)",
    )
    parser.add_argument(
        "--role-arn",
        help=(
            "ARN del rol a asumir para cross-account. Usar {account_id} como "
            "placeholder (ej: arn:aws:iam::{account_id}:role/AuditRole)"
        ),
    )
    parser.add_argument(
        "--regions",
        help=f"Regiones AWS separadas por coma (default: {','.join(DEFAULT_REGIONS)})",
    )
    parser.add_argument(
        "--output",
        default="./hub-spoke-output",
        help="Directorio de salida (default: ./hub-spoke-output)",
    )
    return parser.parse_args()


def main():
    args = parse_args()

    # Banner
    write_status("═══════════════════════════════════════════════════════════════")
    write_status("  AWS Hub & Spoke — Deep Discovery & Best Practices Audit     ")
    write_status(f"  Version: {SCRIPT_VERSION}                                  ")
    write_status("═══════════════════════════════════════════════════════════════")

    # Regiones
    regions = args.regions.split(",") if args.regions else DEFAULT_REGIONS
    regions = [r.strip() for r in regions]
    write_status(f"Regiones objetivo: {', '.join(regions)}")

    # Output
    os.makedirs(args.output, exist_ok=True)
    write_status(f"Directorio de salida: {args.output}")

    # Resolver cuentas
    write_status("\nResolviendo cuentas AWS...")
    accounts = resolve_accounts(args)
    if not accounts:
        write_status("No se resolvieron cuentas. Verificar credenciales.", "ERROR")
        sys.exit(1)
    write_status(f"Cuentas a analizar: {len(accounts)}\n")

    # Data Store
    store = DiscoveryStore()

    # ── PILAR 1: Recolección ──
    for acct in accounts:
        write_status("────────────────────────────────────────────────────────────")
        write_status(f"Procesando cuenta: {acct['label']} [{acct['account_id']}]")
        try:
            collect_resources(
                session=acct["session"],
                account_id=acct["account_id"],
                account_label=acct["label"],
                regions=regions,
                store=store,
            )
        except Exception as e:
            write_status(f"  ERROR procesando cuenta {acct['account_id']}: {e}", "ERROR")
            write_status("  Continuando con la siguiente cuenta...", "WARN")
            traceback.print_exc()
            continue

    # Resumen de recolección
    write_status("\n════════════════════════════════════════════════════════════")
    write_status("Recolección completa. Resumen de recursos:", "OK")
    for metric, value in store.summary().items():
        write_status(f"  {metric}: {value}")

    # ── PILAR 2: Clasificación ──
    classify_hub_spoke(store)

    # ── PILAR 3: Best Practices ──
    evaluate_best_practices(store)

    # ── PILAR 4: Mermaid ──
    mermaid_content = generate_mermaid(store)

    # Mostrar Mermaid en consola
    write_status("\n┌─────────────────────────────────────────────────────────┐")
    write_status("│              MERMAID OUTPUT (copiar abajo)              │")
    write_status("└─────────────────────────────────────────────────────────┘")
    print("")
    print("```mermaid")
    print(mermaid_content)
    print("```")

    # ── PILAR 5: Exportación ──
    files = export_all(store, mermaid_content, args.output)

    # Resumen Final
    write_status("\n═══════════════════════════════════════════════════════════")
    write_status("                   EJECUCIÓN COMPLETA                      ")
    write_status("═══════════════════════════════════════════════════════════")
    write_status("Archivos generados:", "OK")
    if "xlsx" in files:
        write_status(f"  📊 {files['xlsx']}")
    else:
        write_status(f"  📊 {args.output}/inventory-*.csv + best-practices-report.csv")
    write_status(f"  🗺️  {files.get('mermaid', '')}")
    write_status(f"  📝 {files.get('summary_md', '')}")
    write_status("")
    write_status("Próximos pasos:")
    write_status("  1. Abrir el libro Excel y revisar la hoja 'Best Practices'")
    write_status("  2. Pegar el contenido .mmd en https://mermaid.live para visualizar")
    write_status("  3. Integrar en pipeline de gobernanza (AWS Config / CodePipeline)")
    write_status("═══════════════════════════════════════════════════════════")


if __name__ == "__main__":
    main()
